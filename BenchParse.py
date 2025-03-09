# ----- License -------------------------------------------------------- #

# MIT License (c) 2025 Cursed
# Free to use, modify, and distribute. No warranty, no liability.  
# Use responsibly — don’t be a script kiddie.

# ----- Libraries ------------------------------------------------------ #

import re
import os
import pdfplumber
from halo import Halo
from rich.console import Console
from rich.text import Text
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import Workbook

# ----- Global Declaration --------------------------------------------- #

console = Console()
spinner = Halo(color="green")
#recommendation_pattern = re.compile(r'^\s*(\d+(?:\.\d+)+)\s+(.+)')
remove_pattern = re.compile(r'Page\s\d{1,3}|•')
title_pattern = re.compile(r'^(\d+\.\d+(?:\.\d+)*)\s*(\(L\d+\))?\s*(.*)')
page_number_pattern = re.compile(r'\bPage\s+\d+\b', re.IGNORECASE)
sections = {
    'Description': 'Description:',
    'Rationale': 'Rationale:',
    'Impact': 'Impact:',
    'Audit': 'Audit:',
    'Profile Applicability': 'Profile Applicability:',
    'Remediation': 'Remediation:'
}

# ----- Banner --------------------------------------------------------- #

def banner():
    console.print(rf"""[bold yellow]
┌──────────────────────────────────────────────────────────────┐
│                                                              │                                                                                      
│     [bold green]____                  _     ____                         [bold yellow]│                    
│    [bold green]| __ )  ___ _ __   ___| |__ |  _ \ __ _ _ __ ___  ___     [bold yellow]│
│    [bold green]|  _ \ / _ \ '_ \ / __| '_ \| |_) / _` | '__/ __|/ _ \    [bold yellow]│
│    [bold green]| |_) |  __/ | | | (__| | | |  __/ (_| | |  \__ \  __/    [bold yellow]│
│    [bold green]|____/ \___|_| |_|\___|_| |_|_|   \__,_|_|  |___/\___|    [bold yellow]│
│                                                              │
│                                                              │                  
└──────────────────────────────────────────────────────────────┘                                                                        
    """)
    console.print("[bold green]+--------------------------------------------------------------+")
    console.print("[bold green]  BenchParse - CIS Benchmark to Excel Converter")
    console.print("[bold green]  Created by [bold black]Cursed271")
    console.print("[bold green]+--------------------------------------------------------------+")

# ----- Remove Page Numbers -------------------------------------------- #

def remove_pgno(text):
    return page_number_pattern.sub('', text)

# ----- Save to Excel -------------------------------------------------- #

def save_output(recommendations, output_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CIS Benchmark"
    header_font = Font(name = "Aptos", size = 12, bold = True, color = "000000")
    header_fill = PatternFill(start_color = "00B0F0", end_color = "00B0F0", fill_type = "solid")
    cell_font = Font(name = "Aptos", size = 12, color = "000000")
    cell_fill = PatternFill(start_color = "FFFFFF", end_color = "FFFFFF", fill_type = "solid")
    headers = ['Control Name', 'Control Title', 'Description', 'Rationale', 'Impact', 'Audit', 'Recommendation', 'Profile Applicability']
    sheet.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = sheet.cell(row = 1, column = col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal = 'left')
    for recommendation in recommendations:
        row_values = [
            recommendation.get('Number', ''),
            recommendation.get('Title', ''),
            recommendation.get('Description', ''),
            recommendation.get('Rationale', ''),
            recommendation.get('Impact', ''),
            recommendation.get('Audit', ''),
            recommendation.get('Remediation', ''),
            recommendation.get('Profile Applicability', '')
        ]
        sheet.append(row_values)
    for row_num in range(2, len(recommendations) + 2):
        for col_num in range(1, len(headers) + 1):
            cell = sheet.cell(row = row_num, column = col_num)
            cell.font = cell_font
            cell.fill = cell_fill
            cell.alignment = Alignment(horizontal = 'left')
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 35
    sheet.column_dimensions['D'].width = 35
    sheet.column_dimensions['E'].width = 35
    sheet.column_dimensions['F'].width = 35
    sheet.column_dimensions['G'].width = 35
    sheet.column_dimensions['H'].width = 35
    num_rows = len(recommendations)
    num_cols = len(headers)
    tab_range = f"A1:{chr(64 + num_cols)}{num_rows + 1}"
    tab = Table(displayName = "Benchmark", ref = tab_range)
    style = TableStyleInfo(
        name = "TableStyleMedium2", 
        showFirstColumn = False, 
        showLastColumn = False, 
        showRowStripes = False, 
        showColumnStripes = False, 
    )
    tab.tableStyleInfo = style
    sheet.add_table(tab)
    workbook.save(output_path)

# ----- Read CIS Benchmark --------------------------------------------- #

def read_pdf(input_path):
    text = []
    with pdfplumber.open(input_path) as pdf:
        total_pages = len(pdf.pages)
        extraction_started = False
        for page_number, page in enumerate(pdf.pages[9:], start = 10):
            page_text = page.extract_text()
            if not extraction_started:
                if "Recommendations" in page_text and "....." not in page_text and "Recommendation Definitions" not in page_text:
                    extraction_started = True
            if extraction_started:
                if "Appendix: Summary Table" in page_text or "Checklist" in page_text:
                    break
                text.append(page_text)
    return '\n'.join(text)

# ----- Extract Profile Applicability ---------------------------------- #

def extract_profile(lines, start_index, max_depth = 10):
    for i in range(start_index + 1, min(start_index + max_depth, len(lines))):
        line = lines[i].strip()
        if line.startswith("Profile Applicability:"):
            return True
        if title_pattern.match(line) or any(line.startswith(sec) for sec in sections.values()):
            return False
    return False

# ----- Extract Recommendation ----------------------------------------- #

def extract_recommend(text):
    recommendations = []
    lines = text.splitlines()
    current_recommendation = {}
    current_index = 0
    while current_index < len(lines):
        line = lines[current_index].strip()
        line = remove_pgno(line)
        title_match = title_pattern.match(line)
        if title_match:
            if extract_profile(lines, current_index):
                if current_recommendation:
                    recommendations.append(current_recommendation)
                current_recommendation = {
                    'Number': title_match.group(1),
                    'Title': title_match.group(3),
                }
                while (
                    current_index + 1 < len(lines) and 
                    not any (lines[current_index + 1].strip().startswith(sec) for sec in sections.values()) and
                    not title_pattern.match(lines[current_index + 1].strip())
                ):
                    current_index += 1
                    current_recommendation['Title'] += " " + lines[current_index].strip()
        for section_name, section_start in sections.items():
            if line.startswith(section_start):
                content, next_index = extract_section(lines, current_index)
                current_recommendation[section_name] = content
                current_index = next_index - 1
                break
        current_index += 1
    if current_recommendation:
        recommendations.append(current_recommendation)
    unique_recommendations = {(rec['Number'], rec['Title']): rec for rec in recommendations}
    return list(unique_recommendations.values())

# ----- Extract Section ------------------------------------------------ #

def extract_section(lines, start_index):
    content = []
    current_index = start_index + 1
    while current_index < len(lines):
        line = lines[current_index].strip()
        line = remove_pgno(line)
        if any(line.startswith(sec) for sec in sections.values()) or title_pattern.match(line) or 'CIS Controls' in line:
            break
        content.append(line)
        current_index += 1
    return ' '.join(content).strip(), current_index

# ----- Menu ----------------------------------------------------------- #

def menu():
    input_path = console.input("[green][+] Enter the Benchmark that will be converted to Excel: ")
    output_path = console.input("[green][+] Enter the name of the Output File: ")
    spinner.start()
    text = read_pdf(input_path)
    recommendations = extract_recommend(text)
    save_output(recommendations, output_path)
    #spinner.stop()
    console.print(f"[green][+] Finished completed the Benchmark to Excel - {output_path}")
    console.print("[bold green]+--------------------------------------------------------------+")

# ----- Main Function -------------------------------------------------- #

if __name__ == "__main__":
    os.system("cls" if os.name == "nt" else "clear")
    banner()
    menu()

# ----- End ------------------------------------------------------------ #